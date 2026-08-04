#!/usr/bin/env python3
"""Isolated Local AI technical PoC harness — BGE-M3 embedding-first evaluation.

Constraints (hard):
- Does NOT activate app production scoring (mock path stays sole live path).
- Does NOT install or run Qwen.
- Does NOT connect to Preview/Production DB.
- Never prints or writes raw Q/A text, embeddings, model files, or secrets.
- Model/cache/artifacts live under /tmp/unstandard-local-ai/ only.

Usage:
  export UNSTANDARD_LABELING_WORKBOOK_PATH=/absolute/path/to/Unstandard_LabelingDataset_v0.1.xlsx
  /tmp/unstandard-local-ai/venv/bin/python scripts/local-ai/poc_bge_m3.py

Exit codes:
  0  TECHNICAL_POC_PASS or TECHNICAL_POC_CONDITIONAL (report written)
  2  BLOCKED_INPUT_FILE_NOT_FOUND / BLOCKED_INPUT_HASH_MISMATCH / other BLOCKED_*
  3  TECHNICAL_POC_FAIL
"""

from __future__ import annotations

import argparse
import json
import os
import platform
import resource
import shutil
import sys
import time
import traceback
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence

_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

from helpers import (
    APPROVED_SNAPSHOT_ID,
    EXPECTED_EMBEDDING_DIM,
    EXPECTED_PHYSICAL_ROWS,
    EXPECTED_UNIQUE_PAIRS,
    EXPECTED_WORKBOOK_SHA256,
    FAST_TRACK_MIN_LENGTH,
    FAST_TRACK_THRESHOLD,
    HUMAN_LABEL_GATE_STATUS,
    LOCAL_AI_ROOT,
    METRIC_AGREEMENT,
    MIN_ANSWER_LENGTH,
    MODEL_ID,
    QWEN_STATUS,
    DEPTH_GRAY_BAND,
    DEPTH_SCORE_THRESHOLD,
    HashMismatchError,
    agreement_with_synthetic_prior,
    ascii_space_removed_length,
    count_distribution,
    embedding_health,
    max_abs_diff,
    normalize_pair_key,
    percentile,
    score_pair,
    sha256_file,
    threshold_band,
    verify_workbook_hash,
)

LABELING_SHEET_HINTS = ("라벨링 세트", "labeling")
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "id": ("id", "번호", "row_id", "sample_id"),
    "category": ("카테고리", "category", "그룹", "group"),
    "question": ("질문", "question", "question_text", "q"),
    "answer": ("답변", "answer", "answer_text", "a"),
    "answer_length": ("답변 길이", "답변길이", "answer_length", "length"),
    "recommended_label": ("권장 레이블", "권장레이블", "recommended_label", "label"),
    "recommended_path": ("권장 경로", "권장경로", "recommended_path", "path"),
}


@dataclass
class RowRecord:
    category: str
    recommended_label: str
    recommended_path: str
    question: str
    answer: str
    stored_answer_length: int | None
    pair_key: str


class PocBlocked(Exception):
    def __init__(self, code: str, detail: str = "") -> None:
        self.code = code
        self.detail = detail
        super().__init__(code)


def _utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _ensure_dirs() -> dict[str, Path]:
    root = Path(LOCAL_AI_ROOT)
    paths = {
        "root": root,
        "models": root / "models",
        "cache": root / "cache",
        "artifacts": root / "artifacts",
        "reports": root / "reports",
        "venv": root / "venv",
    }
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def _rss_bytes() -> int:
    # Linux: ru_maxrss is kilobytes
    usage = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    if platform.system() == "Darwin":
        return int(usage)
    return int(usage) * 1024


def collect_preflight() -> dict[str, Any]:
    mem_total = None
    mem_available = None
    try:
        with open("/proc/meminfo", "r", encoding="utf-8") as handle:
            for line in handle:
                if line.startswith("MemTotal:"):
                    mem_total = int(line.split()[1]) * 1024
                elif line.startswith("MemAvailable:"):
                    mem_available = int(line.split()[1]) * 1024
    except OSError:
        pass

    disk = shutil.disk_usage(LOCAL_AI_ROOT)
    gpu = {"available": False, "detail": "nvidia-smi_absent"}
    if shutil.which("nvidia-smi"):
        gpu = {"available": True, "detail": "nvidia-smi_present_not_queried"}

    torch_info: dict[str, Any] = {"present": False}
    st_info: dict[str, Any] = {"present": False}
    try:
        import torch  # type: ignore

        torch_info = {
            "present": True,
            "version": getattr(torch, "__version__", "unknown"),
            "cuda_available": bool(torch.cuda.is_available()),
        }
    except Exception:
        torch_info = {"present": False}

    try:
        import sentence_transformers  # type: ignore

        st_info = {
            "present": True,
            "version": getattr(sentence_transformers, "__version__", "unknown"),
        }
    except Exception:
        st_info = {"present": False}

    return {
        "os": platform.platform(),
        "arch": platform.machine(),
        "cpu_count": os.cpu_count(),
        "python": sys.version.split()[0],
        "executable": sys.executable,
        "ram_total_bytes": mem_total,
        "ram_available_bytes": mem_available,
        "disk_free_bytes": disk.free,
        "disk_total_bytes": disk.total,
        "gpu": gpu,
        "torch": torch_info,
        "sentence_transformers": st_info,
        "qwen_status": QWEN_STATUS,
        "local_ai_root": LOCAL_AI_ROOT,
    }


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    # Strip emoji/symbols for matching
    cleaned = "".join(ch for ch in text if ch.isalnum() or ch.isspace() or ch in "_-")
    return " ".join(cleaned.split())


def _map_headers(headers: Sequence[Any]) -> dict[str, int]:
    normalized = [_normalize_header(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, aliases in HEADER_ALIASES.items():
        for idx, header in enumerate(normalized):
            for alias in aliases:
                alias_n = _normalize_header(alias)
                if header == alias_n or alias_n in header:
                    mapping[field] = idx
                    break
            if field in mapping:
                break
    required = ("category", "question", "answer", "recommended_label")
    missing = [name for name in required if name not in mapping]
    if missing:
        raise PocBlocked(
            "BLOCKED_RUNTIME",
            f"workbook_header_unmapped:{','.join(missing)}",
        )
    return mapping


def load_workbook_rows(path: str) -> tuple[list[RowRecord], dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise PocBlocked("BLOCKED_RUNTIME", "openpyxl_missing") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = None
    for name in workbook.sheetnames:
        if any(hint in name for hint in LABELING_SHEET_HINTS):
            sheet = workbook[name]
            break
    if sheet is None:
        # Fall back to first sheet
        sheet = workbook[workbook.sheetnames[0]]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration as exc:
        raise PocBlocked("BLOCKED_RUNTIME", "workbook_empty") from exc

    mapping = _map_headers(header)
    records: list[RowRecord] = []
    for raw in rows_iter:
        if raw is None:
            continue
        question = str(raw[mapping["question"]] or "").strip()
        answer = str(raw[mapping["answer"]] or "").strip()
        if not question and not answer:
            continue
        category = str(raw[mapping["category"]] or "").strip()
        recommended_label = str(raw[mapping["recommended_label"]] or "").strip().upper()
        recommended_path = ""
        if "recommended_path" in mapping:
            recommended_path = str(raw[mapping["recommended_path"]] or "").strip().upper()
        stored_len = None
        if "answer_length" in mapping and raw[mapping["answer_length"]] is not None:
            try:
                stored_len = int(raw[mapping["answer_length"]])
            except (TypeError, ValueError):
                stored_len = None
        records.append(
            RowRecord(
                category=category,
                recommended_label=recommended_label,
                recommended_path=recommended_path,
                question=question,
                answer=answer,
                stored_answer_length=stored_len,
                pair_key=normalize_pair_key(question, answer),
            )
        )
    workbook.close()

    unique_keys = {row.pair_key for row in records}
    length_ok = 0
    length_checked = 0
    for row in records:
        if row.stored_answer_length is None:
            continue
        length_checked += 1
        if row.stored_answer_length == ascii_space_removed_length(row.answer):
            length_ok += 1

    meta = {
        "physical_rows": len(records),
        "unique_pairs": len(unique_keys),
        "sheet_used": sheet.title,
        "category_distribution_physical": count_distribution(r.category for r in records),
        "recommended_label_distribution_physical": count_distribution(
            r.recommended_label for r in records
        ),
        "answer_length_invariant_ok": length_ok,
        "answer_length_invariant_checked": length_checked,
        "expected_physical_rows": EXPECTED_PHYSICAL_ROWS,
        "expected_unique_pairs": EXPECTED_UNIQUE_PAIRS,
    }
    return records, meta


def dedupe_unique_pairs(rows: list[RowRecord]) -> list[RowRecord]:
    """Keep first occurrence of each normalized Q/A pair."""
    seen: set[str] = set()
    unique: list[RowRecord] = []
    for row in rows:
        if row.pair_key in seen:
            continue
        seen.add(row.pair_key)
        unique.append(row)
    return unique


def _set_hf_env(cache_dir: Path) -> None:
    cache = str(cache_dir)
    os.environ.setdefault("HF_HOME", cache)
    os.environ.setdefault("HUGGINGFACE_HUB_CACHE", cache)
    os.environ.setdefault("TRANSFORMERS_CACHE", cache)
    os.environ.setdefault("SENTENCE_TRANSFORMERS_HOME", cache)
    # Keep downloads under the isolated tree; do not write into the repo.
    os.environ.setdefault("HF_HUB_DISABLE_TELEMETRY", "1")


def load_embedder(cache_dir: Path, models_dir: Path) -> tuple[Any, dict[str, Any]]:
    _set_hf_env(cache_dir)
    try:
        from sentence_transformers import SentenceTransformer  # type: ignore
    except ImportError as exc:
        raise PocBlocked(
            "BLOCKED_RUNTIME",
            "sentence_transformers_missing",
        ) from exc

    started = time.perf_counter()
    try:
        model = SentenceTransformer(
            MODEL_ID,
            cache_folder=str(cache_dir),
        )
    except Exception as exc:
        # Do not leak exception text that might include paths with secrets.
        raise PocBlocked("BLOCKED_MODEL_DOWNLOAD", type(exc).__name__) from None

    load_ms = int((time.perf_counter() - started) * 1000)
    revision = None
    model_digest = None
    try:
        from huggingface_hub import snapshot_download  # type: ignore

        # Resolve the cached snapshot without re-downloading when warm.
        snap = snapshot_download(
            repo_id=MODEL_ID,
            cache_dir=str(cache_dir),
            local_files_only=True,
        )
        revision = Path(snap).name
        refs_main = Path(cache_dir) / f"models--{MODEL_ID.replace('/', '--')}" / "refs" / "main"
        if refs_main.is_file():
            revision = refs_main.read_text(encoding="utf-8").strip() or revision
        # Prefer a content digest of a small config artifact over weight dumps.
        config_path = Path(snap) / "config.json"
        if config_path.is_file():
            model_digest = sha256_file(str(config_path))
    except Exception:
        revision = None
        model_digest = None
    if not revision:
        try:
            revision = getattr(getattr(model, "model_card_data", None), "base_model_revision", None)
        except Exception:
            revision = None

    meta = {
        "backend": "sentence-transformers",
        "model_id": MODEL_ID,
        "load_ms": load_ms,
        "cache_dir": str(cache_dir),
        "models_dir": str(models_dir),
        "model_revision": str(revision) if revision else "unresolved",
        "model_config_sha256": model_digest,
        "torch_device": str(getattr(model, "device", "unknown")),
        "expected_embedding_dim": EXPECTED_EMBEDDING_DIM,
    }
    return model, meta


def embed_texts(model: Any, texts: list[str], batch_size: int = 16) -> list[list[float]]:
    # normalize_embeddings=True is BGE-M3 recommended practice for cosine.
    vectors = model.encode(
        texts,
        batch_size=batch_size,
        show_progress_bar=False,
        normalize_embeddings=True,
        convert_to_numpy=True,
    )
    return [vector.astype(float).tolist() for vector in vectors]


def run_poc(
    workbook_path: str,
    out_dir: Path,
    *,
    smoke_physical_latency: bool,
    max_pairs: int | None,
) -> dict[str, Any]:
    paths = _ensure_dirs()

    # Input gates first — file/hash blockers must win over runtime absence.
    if not workbook_path:
        raise PocBlocked("BLOCKED_INPUT_FILE_NOT_FOUND", "UNSTANDARD_LABELING_WORKBOOK_PATH unset")
    if not Path(workbook_path).is_file():
        raise PocBlocked("BLOCKED_INPUT_FILE_NOT_FOUND", "path_missing")

    try:
        observed_hash = verify_workbook_hash(workbook_path, EXPECTED_WORKBOOK_SHA256)
    except HashMismatchError as exc:
        raise PocBlocked("BLOCKED_INPUT_HASH_MISMATCH", f"observed_sha256={exc.observed}") from None

    preflight = collect_preflight()

    # Soft resource gates — fail closed when clearly insufficient.
    if preflight["disk_free_bytes"] is not None and preflight["disk_free_bytes"] < 8 * 1024**3:
        raise PocBlocked("BLOCKED_RESOURCE_LIMIT", "disk_free_lt_8gib")
    if (
        preflight["ram_available_bytes"] is not None
        and preflight["ram_available_bytes"] < 2 * 1024**3
    ):
        raise PocBlocked("BLOCKED_RESOURCE_LIMIT", "ram_available_lt_2gib")
    if not preflight["sentence_transformers"]["present"] or not preflight["torch"]["present"]:
        raise PocBlocked("BLOCKED_RUNTIME", "embedding_runtime_absent")

    rows, workbook_meta = load_workbook_rows(workbook_path)
    unique_rows = dedupe_unique_pairs(rows)
    if max_pairs is not None:
        unique_rows = unique_rows[: max(0, max_pairs)]

    model, model_meta = load_embedder(paths["cache"], paths["models"])

    # Determinism probe on a tiny synthetic phrase (not workbook text).
    probe = ["unstandard local ai determinism probe"]
    rss_before = _rss_bytes()
    t0 = time.perf_counter()
    emb_a = embed_texts(model, probe, batch_size=1)[0]
    cold_ms = (time.perf_counter() - t0) * 1000
    t1 = time.perf_counter()
    emb_b = embed_texts(model, probe, batch_size=1)[0]
    warm_ms = (time.perf_counter() - t1) * 1000
    determinism_max_abs = max_abs_diff(emb_a, emb_b)

    # Optional physical-row latency smoke (does not drive score distribution).
    physical_latency_ms: list[float] = []
    if smoke_physical_latency and rows:
        sample = rows[: min(32, len(rows))]
        for row in sample:
            started = time.perf_counter()
            _ = embed_texts(model, [row.question, row.answer], batch_size=2)
            physical_latency_ms.append((time.perf_counter() - started) * 1000)

    # Unique-pair embedding + scoring (distribution source of truth).
    success = 0
    fail = 0
    latencies_ms: list[float] = []
    scores: list[float] = []
    verdicts: list[str] = []
    paths_out: list[str] = []
    categories: list[str] = []
    synthetic_labels: list[str] = []
    bands: list[str] = []
    all_answer_embeddings: list[list[float]] = []

    for row in unique_rows:
        try:
            started = time.perf_counter()
            q_emb, a_emb = embed_texts(model, [row.question, row.answer], batch_size=2)
            elapsed = (time.perf_counter() - started) * 1000
            result = score_pair(row.question, row.answer, q_emb, a_emb)
            latencies_ms.append(elapsed)
            scores.append(float(result["depth_score"]))
            verdicts.append(str(result["verdict"]))
            paths_out.append(str(result["path"]))
            categories.append(row.category)
            synthetic_labels.append(row.recommended_label)
            bands.append(threshold_band(float(result["depth_score"])))
            all_answer_embeddings.append(a_emb)
            success += 1
        except Exception:
            fail += 1

    rss_after = _rss_bytes()
    latencies_sorted = sorted(latencies_ms)
    total_embed_s = sum(latencies_ms) / 1000.0 if latencies_ms else 0.0
    throughput = (success / total_embed_s) if total_embed_s > 0 else None

    agreement = agreement_with_synthetic_prior(verdicts, synthetic_labels)
    health = embedding_health(all_answer_embeddings)

    # Score distribution by category (unique pairs only).
    by_category: dict[str, dict[str, Any]] = {}
    for category, score, verdict in zip(categories, scores, verdicts):
        bucket = by_category.setdefault(
            category,
            {"n": 0, "scores": [], "verdict_counts": Counter()},
        )
        bucket["n"] += 1
        bucket["scores"].append(score)
        bucket["verdict_counts"][verdict] += 1

    category_summary: dict[str, Any] = {}
    for category, bucket in sorted(by_category.items()):
        cat_scores = sorted(bucket["scores"])
        category_summary[category] = {
            "n": bucket["n"],
            "score_p50": percentile(cat_scores, 50),
            "score_p95": percentile(cat_scores, 95),
            "score_mean": round(sum(cat_scores) / len(cat_scores), 6) if cat_scores else None,
            "verdict_counts": dict(bucket["verdict_counts"]),
        }

    soft_issues: list[str] = []
    hard_issues: list[str] = []

    if success == 0:
        hard_issues.append("zero_successful_pairs")
    if fail > 0:
        soft_issues.append("pair_failures_present")
    if not health["dim_ok"]:
        hard_issues.append("embedding_dim_mismatch")
    if health["nan_count"] or health["inf_count"]:
        hard_issues.append("nan_or_inf_in_embeddings")
    if determinism_max_abs > 1e-5:
        soft_issues.append("determinism_drift")
    p95 = percentile(latencies_sorted, 95)
    if p95 is not None and p95 > 1200:
        soft_issues.append("p95_latency_gt_1200ms")
    if workbook_meta["physical_rows"] != EXPECTED_PHYSICAL_ROWS:
        soft_issues.append("physical_row_count_unexpected")
    if workbook_meta["unique_pairs"] != EXPECTED_UNIQUE_PAIRS and max_pairs is None:
        soft_issues.append("unique_pair_count_unexpected")

    if hard_issues:
        verdict = "TECHNICAL_POC_FAIL"
    elif soft_issues:
        verdict = "TECHNICAL_POC_CONDITIONAL"
    else:
        verdict = "TECHNICAL_POC_PASS"

    report: dict[str, Any] = {
        "schema_version": "local-ai-technical-poc-v1",
        "generated_at_utc": _utc_now(),
        "verdict": verdict,
        "alpha_readiness": "NOT_CLAIMED",
        "production_scoring": "UNTOUCHED_MOCK_ACTIVE",
        "human_label_gate": HUMAN_LABEL_GATE_STATUS,
        "qwen_status": QWEN_STATUS,
        "snapshot": {
            "approved_snapshot_id": APPROVED_SNAPSHOT_ID,
            "workbook_sha256": observed_hash,
            "expected_sha256": EXPECTED_WORKBOOK_SHA256,
            "hash_matched": True,
        },
        "canonical_thresholds": {
            "min_answer_length": MIN_ANSWER_LENGTH,
            "fast_track_min_length": FAST_TRACK_MIN_LENGTH,
            "fast_track_threshold": FAST_TRACK_THRESHOLD,
            "threshold": DEPTH_SCORE_THRESHOLD,
            "gray_band": DEPTH_GRAY_BAND,
        },
        "preflight": preflight,
        "model": model_meta,
        "dataset": {
            **workbook_meta,
            "scoring_universe": "unique_pairs",
            "unique_pairs_scored": success,
            "unique_pair_failures": fail,
            "max_pairs_cap": max_pairs,
        },
        "embedding_health": health,
        "determinism": {
            "probe": "synthetic_fixed_phrase",
            "max_abs_diff": determinism_max_abs,
            "cold_latency_ms": round(cold_ms, 3),
            "warm_latency_ms": round(warm_ms, 3),
        },
        "latency_unique_pairs_ms": {
            "n": len(latencies_sorted),
            "p50": percentile(latencies_sorted, 50),
            "p95": p95,
            "mean": round(sum(latencies_sorted) / len(latencies_sorted), 3)
            if latencies_sorted
            else None,
        },
        "latency_physical_smoke_ms": {
            "enabled": smoke_physical_latency,
            "n": len(physical_latency_ms),
            "p50": percentile(sorted(physical_latency_ms), 50) if physical_latency_ms else None,
            "p95": percentile(sorted(physical_latency_ms), 95) if physical_latency_ms else None,
            "note": "smoke only; score distribution uses unique pairs",
        },
        "resources": {
            "rss_before_bytes": rss_before,
            "rss_after_bytes": rss_after,
            "rss_delta_bytes": rss_after - rss_before,
            "peak_rss_bytes": rss_after,
            "throughput_pairs_per_sec": throughput,
        },
        "score_distribution_unique_pairs": {
            "n": len(scores),
            "p50": percentile(sorted(scores), 50) if scores else None,
            "p95": percentile(sorted(scores), 95) if scores else None,
            "mean": round(sum(scores) / len(scores), 6) if scores else None,
            "threshold_band_counts": count_distribution(bands),
            "verdict_counts": count_distribution(verdicts),
            "path_counts": count_distribution(paths_out),
            "by_category": category_summary,
        },
        METRIC_AGREEMENT: agreement,
        "issues": {"hard": hard_issues, "soft": soft_issues},
        "redaction": {
            "raw_qa_text": "omitted",
            "embeddings": "omitted",
            "pair_keys": "omitted",
            "secrets": "omitted",
        },
        "notes": [
            "Technical PoC only. Not calibration. Not alpha readiness.",
            "agreement_with_synthetic_prior is an offline synthetic-label proxy, not human accuracy.",
            "Qwen remains INACTIVE_NOT_INSTALLED for this harness.",
        ],
    }

    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_path = out_dir / f"bge_m3_technical_poc_{stamp}.json"
    with open(out_path, "w", encoding="utf-8") as handle:
        json.dump(report, handle, ensure_ascii=True, indent=2, sort_keys=True)
        handle.write("\n")
    report["report_path"] = str(out_path)
    return report


def _print_safe_summary(report: dict[str, Any]) -> None:
    # Aggregate-only. Never print workbook text or embeddings.
    print(f"verdict={report.get('verdict')}")
    print(f"report_path={report.get('report_path')}")
    dataset = report.get("dataset") or {}
    print(
        "rows="
        f"physical:{dataset.get('physical_rows')} "
        f"unique:{dataset.get('unique_pairs')} "
        f"scored:{dataset.get('unique_pairs_scored')}"
    )
    print(f"human_label_gate={report.get('human_label_gate')}")
    print(f"qwen_status={report.get('qwen_status')}")
    agreement = report.get(METRIC_AGREEMENT) or {}
    if agreement.get("rate") is not None:
        print(f"{METRIC_AGREEMENT}={agreement.get('rate')} n={agreement.get('n')}")
    issues = report.get("issues") or {}
    if issues.get("hard") or issues.get("soft"):
        print(f"issues_hard={issues.get('hard')}")
        print(f"issues_soft={issues.get('soft')}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Isolated BGE-M3 Local AI technical PoC (redacted aggregates only)."
    )
    parser.add_argument(
        "--workbook",
        default=os.environ.get("UNSTANDARD_LABELING_WORKBOOK_PATH", ""),
        help="Absolute path to labeling workbook (or set UNSTANDARD_LABELING_WORKBOOK_PATH).",
    )
    parser.add_argument(
        "--out",
        default=os.environ.get(
            "UNSTANDARD_LOCAL_AI_POC_OUT",
            str(Path(LOCAL_AI_ROOT) / "reports"),
        ),
        help="Directory for redacted JSON report.",
    )
    parser.add_argument(
        "--smoke-physical-latency",
        action="store_true",
        help="Optionally embed a small physical-row sample for latency smoke only.",
    )
    parser.add_argument(
        "--max-pairs",
        type=int,
        default=None,
        help="Cap unique pairs scored (debug only; full run omits this).",
    )
    parser.add_argument(
        "--preflight-only",
        action="store_true",
        help="Print preflight JSON and exit without loading model or workbook.",
    )
    args = parser.parse_args(argv)

    _ensure_dirs()

    if args.preflight_only:
        print(json.dumps(collect_preflight(), indent=2, sort_keys=True))
        return 0

    try:
        report = run_poc(
            args.workbook,
            Path(args.out),
            smoke_physical_latency=bool(args.smoke_physical_latency),
            max_pairs=args.max_pairs,
        )
    except PocBlocked as blocked:
        payload = {
            "verdict": blocked.code,
            "detail": blocked.detail,
            "human_label_gate": HUMAN_LABEL_GATE_STATUS,
            "qwen_status": QWEN_STATUS,
            "alpha_readiness": "NOT_CLAIMED",
            "generated_at_utc": _utc_now(),
        }
        out_dir = Path(args.out)
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        out_path = out_dir / f"bge_m3_technical_poc_{blocked.code}_{stamp}.json"
        with open(out_path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=True, indent=2, sort_keys=True)
            handle.write("\n")
        print(f"verdict={blocked.code}")
        print(f"report_path={out_path}")
        if blocked.detail:
            print(f"detail={blocked.detail}")
        return 2
    except Exception:
        # Redacted failure — no traceback content that could include cell text.
        print("verdict=BLOCKED_RUNTIME")
        print(f"detail={type(sys.exc_info()[1]).__name__}")
        traceback.print_exc(file=sys.stderr)
        return 2

    _print_safe_summary(report)
    if report["verdict"] == "TECHNICAL_POC_FAIL":
        return 3
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
